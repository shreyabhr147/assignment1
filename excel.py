import os
import openpyxl
current_dir=os.path.dirname(__file__)
filename=os.path.join(current_dir,"ass.xlsx")
wb=openpyxl.Workbook()
print("Sheets",wb.sheetnames)
sheet=wb.active
print("Sheet",sheet)
print("Sheet Title",sheet.title)
sheet.title='ML'
print("Sheets",wb.sheetnames)
wb.save(filename)

for i in range(1,101):
    for j in range(65,91):
        a=chr(j)
        z=a+str(i)
        sheet[z]=z

wb.save(filename)
